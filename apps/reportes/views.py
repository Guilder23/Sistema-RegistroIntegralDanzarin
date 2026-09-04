import io
from datetime import datetime
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Q
from django.http import FileResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from apps.danzarines.models import Danzarin
from apps.souvenirs.models import Souvenir
from apps.core.permissions import scope_danzarines, is_administrative


def aplicar_filtros_danzarines(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()
    ciudad = request.GET.get('ciudad', '').strip()
    recibio_souvenir = request.GET.get('recibio_souvenir', '')
    souvenir_id = request.GET.get('souvenir_id', '').strip()
    desde = request.GET.get('desde', '').strip()
    hasta = request.GET.get('hasta', '').strip()
    orden = request.GET.get('orden', 'recientes')

    danzarines = scope_danzarines(Danzarin.objects.select_related('user').all(), request.user)
    if q:
        danzarines = danzarines.filter(
            Q(nombre__icontains=q)
            | Q(apellido__icontains=q)
            | Q(email__icontains=q)
            | Q(carnet_ci__icontains=q)
            | Q(carnet_complemento__icontains=q)
            | Q(ciudad__icontains=q)
            | Q(direccion__icontains=q)
        )
    if estado:
        danzarines = danzarines.filter(membresias__estado=estado).distinct()
    if ciudad:
        danzarines = danzarines.filter(ciudad__icontains=ciudad)
    if desde:
        danzarines = danzarines.filter(fecha_ingreso__gte=desde)
    if hasta:
        danzarines = danzarines.filter(fecha_ingreso__lte=hasta)
    if recibio_souvenir == 'si':
        danzarines = danzarines.filter(recibio_souvenir=True)
    elif recibio_souvenir == 'no':
        danzarines = danzarines.filter(recibio_souvenir=False)
    if souvenir_id:
        danzarines = danzarines.filter(entregas_souvenir__souvenir_id=souvenir_id).distinct()

    ordenamiento = '-fecha_ingreso' if orden == 'recientes' else 'fecha_ingreso'
    danzarines = danzarines.order_by(ordenamiento)

    return danzarines, {
        'q': q,
        'estado': estado,
        'ciudad': ciudad,
        'recibio_souvenir': recibio_souvenir,
        'souvenir_id': souvenir_id,
        'desde': desde,
        'hasta': hasta,
        'orden': orden,
    }


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def reportes_danzarines(request):
    danzarines, filtros = aplicar_filtros_danzarines(request)
    souvenirs = Souvenir.objects.filter(activo=True).order_by('nombre')

    # Agregar nombre del souvenir específico a cada danzarin si hay filtro
    for danzarin in danzarines:
        danzarin.souvenir_recibido_nombre = '-'
        if filtros['souvenir_id']:
            try:
                entrega = danzarin.entregas_souvenir.filter(souvenir_id=filtros['souvenir_id']).first()
                if entrega and entrega.souvenir:
                    danzarin.souvenir_recibido_nombre = entrega.souvenir.nombre
            except:
                pass

    return render(request, 'reportes/reportes_danzarines.html', {
        'danzarines': danzarines,
        'souvenirs': souvenirs,
        **filtros,
    })


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def descargar_reporte_danzarines(request):
    danzarines, filtros = aplicar_filtros_danzarines(request)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Danzarines'

    title_font = Font(bold=True, size=16, color='0B3D91')
    subtitle_font = Font(size=11)
    info_font = Font(size=10)

    sheet['A1'] = 'Club carnaval Oruro - Reporte de Danzarines'
    sheet['A1'].font = title_font
    sheet['A2'] = f'Reporte generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    sheet['A2'].font = subtitle_font
    sheet['A3'] = f'Reportado por: {request.user.get_full_name() or request.user.username}'
    sheet['A3'].font = subtitle_font

    filtros_text = []
    if filtros['q']:
        filtros_text.append(f'Búsqueda: {filtros["q"]}')
    if filtros['estado']:
        filtros_text.append(f'Estado: {filtros["estado"].capitalize()}')
    if filtros['ciudad']:
        filtros_text.append(f'Ciudad: {filtros["ciudad"]}')
    if filtros['desde']:
        filtros_text.append(f'Desde: {filtros["desde"]}')
    if filtros['hasta']:
        filtros_text.append(f'Hasta: {filtros["hasta"]}')
    if filtros['recibio_souvenir']:
        filtros_text.append(f'Recibió souvenir: {filtros["recibio_souvenir"].capitalize()}')
    if filtros['souvenir_id']:
        try:
            souvenir = Souvenir.objects.get(id=filtros['souvenir_id'])
            filtros_text.append(f'Souvenir: {souvenir.nombre}')
        except Souvenir.DoesNotExist:
            pass
    if filtros['orden']:
        orden_label = 'Más recientes' if filtros['orden'] == 'recientes' else 'Más antiguos'
        filtros_text.append(f'Orden: {orden_label}')

    if filtros_text:
        sheet['A4'] = 'Filtros aplicados:'
        sheet['A4'].font = info_font
        sheet['A5'] = ' | '.join(filtros_text)
        sheet['A5'].font = info_font
        start_row = 7
    else:
        start_row = 5

    headers = ['N°', 'Código', 'Danzarin', 'Correo', 'N° Carnet', 'Ciudad', 'Estado', 'Souvenir', 'Souvenir recibido', 'Ingreso']
    sheet.cell(row=start_row, column=1, value='N°')
    sheet.cell(row=start_row, column=2, value='Código')
    sheet.cell(row=start_row, column=3, value='Danzarin')
    sheet.cell(row=start_row, column=4, value='Correo')
    sheet.cell(row=start_row, column=5, value='N° Carnet')
    sheet.cell(row=start_row, column=6, value='Ciudad')
    sheet.cell(row=start_row, column=7, value='Estado')
    sheet.cell(row=start_row, column=8, value='Souvenir')
    sheet.cell(row=start_row, column=9, value='Souvenir recibido')
    sheet.cell(row=start_row, column=10, value='Ingreso')

    header_fill = PatternFill('solid', fgColor='0B3D91')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in sheet[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    current_row = start_row + 1
    for idx, danzarin in enumerate(danzarines, 1):
        sheet.cell(row=current_row, column=1, value=idx)
        sheet.cell(row=current_row, column=2, value=danzarin.codigo_danzarin or '-')
        nombre_completo = f"{danzarin.nombre} {danzarin.apellido_paterno or ''} {danzarin.apellido_materno or ''}".strip().upper()
        sheet.cell(row=current_row, column=3, value=nombre_completo)
        sheet.cell(row=current_row, column=4, value=danzarin.email)
        carnet_completo = f"{danzarin.carnet_ci or ''}{danzarin.carnet_complemento or ''}".strip()
        sheet.cell(row=current_row, column=5, value=carnet_completo or '-')
        sheet.cell(row=current_row, column=6, value=danzarin.ciudad or '-')
        sheet.cell(row=current_row, column=7, value=danzarin.get_estado_display())
        sheet.cell(row=current_row, column=8, value='Sí' if danzarin.recibio_souvenir else 'No')
        # Obtener nombre del souvenir específico si hay filtro
        souvenir_nombre = '-'
        if filtros['souvenir_id']:
            try:
                entrega = danzarin.entregas_souvenir.filter(souvenir_id=filtros['souvenir_id']).first()
                if entrega and entrega.souvenir:
                    souvenir_nombre = entrega.souvenir.nombre
            except:
                pass
        sheet.cell(row=current_row, column=9, value=souvenir_nombre)
        sheet.cell(row=current_row, column=10, value=danzarin.fecha_ingreso.strftime('%d/%m/%Y'))
        current_row += 1

    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 28
    sheet.column_dimensions['D'].width = 32
    sheet.column_dimensions['E'].width = 18
    sheet.column_dimensions['F'].width = 18
    sheet.column_dimensions['G'].width = 16
    sheet.column_dimensions['H'].width = 14
    sheet.column_dimensions['I'].width = 16
    sheet.column_dimensions['J'].width = 20

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top')

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = FileResponse(buffer, as_attachment=True, filename='reporte_danzarines.xlsx')
    response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response['Content-Disposition'] = 'attachment; filename=reporte_danzarines.xlsx'
    return response


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def descargar_reporte_danzarines_pdf(request):
    danzarines, filtros = aplicar_filtros_danzarines(request)
    buffer = io.BytesIO()
    documento = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    estilos = getSampleStyleSheet()
    filas = [['Código', 'Danzarin', 'Correo', 'CI / Carnet', 'Ciudad', 'Estado', 'Pago']]
    for danzarin in danzarines:
        membresia = danzarin.membresias.filter(estado__in=['activo', 'suspendido', 'castigado']).select_related('asociacion', 'conjunto').first()
        filas.append([
            danzarin.codigo_danzarin or '-',
            str(danzarin),
            danzarin.email,
            f'{danzarin.carnet_ci} {danzarin.carnet_complemento}'.strip() or '-',
            danzarin.ciudad or '-',
            membresia.get_estado_display() if membresia else 'Dado de baja',
            membresia.get_estado_pago_display() if membresia else '-',
        ])
    elementos = [Paragraph('Reporte de integrantes', estilos['Title']), Spacer(1, 12)]
    tabla = Table(filas, repeatRows=1)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0B3D91')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#D9DEE7')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elementos.append(tabla)
    documento.build(elementos)
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='reporte_danzarines.pdf', content_type='application/pdf')
