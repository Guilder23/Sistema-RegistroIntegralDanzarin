from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import get_object_or_404, redirect, render

from .models import Socio, Membresia
from .models import UserProfile
from .models import generar_codigo_socio
from django.contrib.auth import update_session_auth_hash
import csv
from io import TextIOWrapper
from django.contrib.auth.models import User
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from apps.core.permissions import scope_socios, is_administrative, can_register_members, get_role, registrar_auditoria, can_manage_users
from apps.core.models import Asociacion, Conjunto
from apps.bloques.models import Bloque


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def listar_socios(request):
    q = request.GET.get('q', '').strip()
    estado = request.GET.get('estado', '').strip()

    socios = scope_socios(Socio.objects.select_related('user'), request.user)
    if q:
        socios = socios.filter(
            Q(nombre__icontains=q)
            | Q(apellido__icontains=q)
            | Q(apellido_paterno__icontains=q)
            | Q(apellido_materno__icontains=q)
            | Q(email__icontains=q)
            | Q(user__username__icontains=q)
            | Q(carnet_ci__icontains=q)
        )
    if estado:
        socios = socios.filter(membresias__estado=estado).distinct()

    paginator = Paginator(socios.order_by('-fecha_ingreso'), 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Asegurar que cada usuario listado tenga un UserProfile para evitar errores en plantillas
    for s in page_obj.object_list:
        s.membresia_visible = s.membresias.filter(estado__in=['activo', 'suspendido', 'castigado']).select_related('asociacion', 'conjunto', 'bloque').first() or s.membresias.order_by('-fecha_ingreso').first()
        try:
            UserProfile.objects.get_or_create(user=s.user)
        except Exception:
            pass

    role = get_role(request.user)
    asociaciones = Asociacion.objects.filter(activo=True)
    conjuntos = Conjunto.objects.filter(activo=True).select_related('asociacion')
    bloques = Bloque.objects.filter(activo=True).select_related('conjunto')
    if role == 'administrador_asociacion':
        asociaciones = asociaciones.filter(pk=request.user.userprofile.asociacion_id)
        conjuntos = conjuntos.filter(asociacion_id=request.user.userprofile.asociacion_id)
        bloques = bloques.filter(conjunto__asociacion_id=request.user.userprofile.asociacion_id)
    elif role == 'administrador_conjunto':
        asociaciones = asociaciones.filter(pk=request.user.userprofile.asociacion_id)
        conjuntos = conjuntos.filter(pk=request.user.userprofile.conjunto_id)
        bloques = bloques.filter(conjunto_id=request.user.userprofile.conjunto_id)

    return render(request, 'socios/socios.html', {
        'page_obj': page_obj,
        'q': q,
        'estado': estado,
        'is_admin': can_register_members(request.user),
        'asociaciones': asociaciones,
        'conjuntos': conjuntos,
        'bloques': bloques,
        'role': role,
    })


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def crear_socio(request):
    if request.method != 'POST':
        return redirect('socios:listar_socios')

    username = request.POST.get('username', '').strip()
    nombre = request.POST.get('nombre', '').strip()
    apellido_paterno = request.POST.get('apellido_paterno', '').strip()
    apellido_materno = request.POST.get('apellido_materno', '').strip()
    apellido = request.POST.get('apellido', '').strip() or f"{apellido_paterno} {apellido_materno}".strip()
    email = request.POST.get('email', '').strip()
    telefono = request.POST.get('telefono', '').strip()
    ciudad = request.POST.get('ciudad', '').strip()
    direccion = request.POST.get('direccion', '').strip()
    password = request.POST.get('password', '')
    fecha_nacimiento = request.POST.get('fecha_nacimiento', '').strip() or None
    razon = request.POST.get('razon', '').strip()
    carnet_ci = request.POST.get('carnet_ci', '').strip()
    carnet_complemento = request.POST.get('carnet_complemento', '').strip()
    observacion = request.POST.get('observacion', '').strip()
    sexo = request.POST.get('sexo', '').strip()
    modalidad = request.POST.get('modalidad', '').strip()

    if not username or not nombre or not (apellido_paterno or apellido) or not email or not password:
        messages.error(request, 'Completa los campos obligatorios.')
        return redirect('socios:listar_socios')

    if User.objects.filter(username=username).exists():
        messages.error(request, 'El nombre de usuario ya existe.')
        return redirect('socios:listar_socios')

    role = get_role(request.user)
    if role == 'administrador_conjunto':
        asociacion = request.user.userprofile.asociacion
        conjunto = request.user.userprofile.conjunto
    elif role == 'administrador_asociacion':
        asociacion = request.user.userprofile.asociacion
        conjunto = Conjunto.objects.filter(pk=request.POST.get('conjunto_id'), asociacion=asociacion, activo=True).first()
    else:
        asociacion_id = request.POST.get('asociacion_id') or None
        asociacion = Asociacion.objects.filter(pk=asociacion_id, activo=True).first() if asociacion_id else None
        conjunto = Conjunto.objects.filter(pk=request.POST.get('conjunto_id'), asociacion=asociacion, activo=True).first() if asociacion else None
    bloque = Bloque.objects.filter(pk=request.POST.get('bloque_id'), conjunto=conjunto, activo=True).first() if conjunto else None
    if not asociacion or not conjunto or not bloque:
        messages.error(request, 'Selecciona una asociación, conjunto y bloque válidos.')
        return redirect('socios:listar_socios')

    user = User.objects.create_user(username=username, email=email, password=password)
    user.first_name = nombre
    user.last_name = apellido_paterno or apellido
    user.save()
    socio = Socio.objects.create(
        user=user,
        codigo_socio=generar_codigo_socio(),
        nombre=nombre,
        apellido_paterno=apellido_paterno,
        apellido_materno=apellido_materno,
        apellido=apellido,
        email=email,
        telefono=telefono,
        ciudad=ciudad,
        direccion=direccion,
        fecha_nacimiento=fecha_nacimiento,
        razon=razon,
        carnet_ci=carnet_ci,
        carnet_complemento=carnet_complemento,
        observacion=observacion,
        sexo=sexo,
        modalidad=modalidad,
        creado_por=request.user,
    )
    from .models import Membresia
    Membresia.inscribir(socio, asociacion, conjunto, bloque, estado_pago='al_dia')
    registrar_auditoria(request.user, 'registro_socio', f'Socio {socio.pk}', nuevo={'socio': socio.pk, 'asociacion': asociacion.pk, 'conjunto': conjunto.pk, 'bloque': bloque.pk}, asociacion=asociacion, conjunto=conjunto)
    messages.success(request, 'Socio registrado correctamente.')
    return redirect('socios:listar_socios')


@login_required
def perfil_socio(request):
    user = request.user
    profile, _ = UserProfile.objects.get_or_create(user=user)
    socio = None
    try:
        socio = user.socio_profile
    except Socio.DoesNotExist:
        socio = None

    entregas = socio.entregas_souvenir.select_related('entregado_por', 'souvenir', 'evento').all() if socio else []
    paginator = Paginator(entregas, 10)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'socios/perfil.html', {
        'socio': socio,
        'page_obj': page_obj,
        'is_admin': request.user.is_staff,
        'user_profile': profile,
    })


@login_required
def editar_perfil(request):
    if request.method != 'POST':
        return redirect('socios:perfil_socio')

    user = request.user
    email = request.POST.get('email', '').strip()
    telefono = request.POST.get('telefono', '').strip()

    # Actualizar email en User
    if email:
        user.email = email
        user.save()

    # Actualizar datos en Socio si existe
    try:
        socio = user.socio_profile
        socio.email = email or socio.email
        socio.telefono = telefono or socio.telefono
        socio.save()
    except Socio.DoesNotExist:
        pass

    messages.success(request, 'Datos de perfil actualizados.')
    return redirect('socios:perfil_socio')


@login_required
def cambiar_contrasena(request):
    if request.method != 'POST':
        return redirect('socios:perfil_socio')

    user = request.user
    current = request.POST.get('current_password', '')
    new1 = request.POST.get('new_password1', '')
    new2 = request.POST.get('new_password2', '')

    # Validar que los tres campos estÃ©n presentes
    if not current or not new1 or not new2:
        messages.error(request, 'Completa los 3 campos requeridos para cambiar la contraseÃ±a.')
        return redirect('socios:perfil_socio')

    if not user.check_password(current):
        messages.error(request, 'La contraseÃ±a actual es incorrecta.')
        return redirect('socios:perfil_socio')

    if new1 != new2:
        messages.error(request, 'Las nuevas contraseÃ±as no coinciden.')
        return redirect('socios:perfil_socio')

    try:
        user.set_password(new1)
        user.save()
        # Mantener la sesiÃ³n activa
        update_session_auth_hash(request, user)
        messages.success(request, 'ContraseÃ±a actualizada correctamente.')
    except Exception:
        messages.error(request, 'No se pudo actualizar la contraseÃ±a.')

    return redirect('socios:perfil_socio')


@login_required
def subir_foto(request):
    if request.method != 'POST':
        return redirect('socios:perfil_socio')

    foto = request.FILES.get('foto')
    user_id = request.POST.get('user_id')

    # Si el usuario es admin puede subir foto para otro usuario
    if user_id and request.user.is_staff:
        from django.contrib.auth.models import User
        target = User.objects.filter(id=user_id).first()
        if not target:
            messages.error(request, 'Usuario no encontrado.')
            return redirect('socios:perfil_socio')
        profile, _ = UserProfile.objects.get_or_create(user=target)
    else:
        profile, _ = UserProfile.objects.get_or_create(user=request.user)

    if foto:
        profile.foto = foto
        profile.save()
        messages.success(request, 'Foto de perfil actualizada.')
    else:
        messages.error(request, 'No se recibiÃ³ archivo.')

    return redirect('socios:perfil_socio')


@login_required
@user_passes_test(can_manage_users, login_url='/login/')
def crear_admin(request):
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        rol = request.POST.get('rol', 'administrador_asociacion')
        asociacion_id = request.POST.get('asociacion_id') or None
        asociacion = Asociacion.objects.filter(pk=asociacion_id, activo=True).first() if asociacion_id else None
        conjunto_id = request.POST.get('conjunto_id') or None
        conjunto = Conjunto.objects.filter(pk=conjunto_id, asociacion=asociacion, activo=True).first() if asociacion and conjunto_id else None
        if rol not in {'superadministrador', 'administrador_asociacion', 'administrador_conjunto'}:
            messages.error(request, 'Selecciona un tipo de administrador válido.')
            return redirect('socios:listar_admins')
        if not username or not email or not password:
            messages.error(request, 'Completa los campos obligatorios.')
            return redirect('socios:crear_admin')
        if rol == 'administrador_asociacion' and not asociacion:
            messages.error(request, 'El Administrador de Asociación debe tener una asociación asignada.')
            return redirect('socios:listar_admins')
        if rol == 'administrador_conjunto' and (not asociacion or not conjunto):
            messages.error(request, 'El Administrador de Conjunto debe tener asociación y conjunto asignados.')
            return redirect('socios:listar_admins')
        if User.objects.filter(username=username).exists():
            messages.error(request, 'El nombre de usuario ya existe.')
            return redirect('socios:crear_admin')
        user = User.objects.create_user(username=username, email=email, password=password)
        user.first_name = first_name
        user.last_name = last_name
        user.is_staff = True
        user.is_superuser = rol == 'superadministrador'
        user.save()
        UserProfile.objects.filter(user=user).update(rol=rol, asociacion=asociacion, conjunto=conjunto)
        registrar_auditoria(
            request.user,
            'creacion_admin',
            f'Admin {user.username} ({rol})',
            nuevo={'username': user.username, 'rol': rol, 'asociacion': asociacion.nombre if asociacion else None, 'conjunto': conjunto.nombre if conjunto else None},
            asociacion=asociacion,
            conjunto=conjunto,
        )
        messages.success(request, 'Administrador creado correctamente.')
        return redirect('socios:listar_admins')
    return redirect('socios:listar_admins')


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def importar_socios(request):
    if request.method == 'POST':
        f = request.FILES.get('file')
        if not f:
            messages.error(request, 'Sube un archivo CSV.')
            return redirect('socios:listar_socios')
        try:
            text = TextIOWrapper(f.file, encoding='utf-8')
            reader = csv.DictReader(text)
            created = 0
            for row in reader:
                username = row.get('username') or row.get('usuario') or ''
                nombre = row.get('nombre') or ''
                apellido_paterno = row.get('apellido_paterno') or row.get('apellido') or ''
                apellido_materno = row.get('apellido_materno') or ''
                apellido = row.get('apellido') or f"{apellido_paterno} {apellido_materno}".strip()
                email = row.get('email') or ''
                password = row.get('password') or User.objects.make_random_password()
                telefono = row.get('telefono') or ''
                ciudad = row.get('ciudad') or ''
                direccion = row.get('direccion') or ''
                fecha_nacimiento = row.get('fecha_nacimiento') or None
                razon = row.get('razon') or ''
                carnet_ci = row.get('carnet_ci') or ''
                carnet_complemento = row.get('carnet_complemento') or ''
                if not username or User.objects.filter(username=username).exists():
                    continue
                user = User.objects.create_user(username=username, email=email, password=password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                Socio.objects.create(user=user, codigo_socio=generar_codigo_socio(), nombre=nombre, apellido_paterno=apellido_paterno, apellido_materno=apellido_materno, apellido=apellido, email=email, telefono=telefono, ciudad=ciudad, direccion=direccion, fecha_nacimiento=fecha_nacimiento, razon=razon, carnet_ci=carnet_ci, carnet_complemento=carnet_complemento)
                created += 1
            messages.success(request, f'Socios importados: {created}')
        except Exception as e:
            messages.error(request, f'Error al procesar el archivo: {e}')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def importar_socios_masivo(request):
    return render(request, 'socios/importar_masivo.html')


def validar_filas_importacion(rows, user):
    role = get_role(user)
    errores = []
    for fila, row in enumerate(rows, start=2):
        valores = list(row) + [''] * max(0, 17 - len(row))
        username, nombre = str(valores[0]).strip(), str(valores[1]).strip()
        asociacion_nombre = str(valores[14]).strip()
        conjunto_nombre = str(valores[15]).strip()
        bloque_nombre = str(valores[16]).strip()
        asociacion = Asociacion.objects.filter(nombre__iexact=asociacion_nombre, activo=True).first()
        conjunto = Conjunto.objects.filter(nombre__iexact=conjunto_nombre, asociacion=asociacion, activo=True).first() if asociacion else None
        bloque = Bloque.objects.filter(nombre__iexact=bloque_nombre, conjunto=conjunto, activo=True).first() if conjunto else None
        fila_errores = []
        if not username: fila_errores.append('falta username')
        if not nombre: fila_errores.append('falta nombre')
        if not asociacion: fila_errores.append(f'asociación "{asociacion_nombre}" no encontrada')
        elif role == 'administrador_asociacion' and asociacion.pk != user.userprofile.asociacion_id: fila_errores.append('asociación fuera de tu ámbito')
        if not conjunto: fila_errores.append(f'conjunto "{conjunto_nombre}" no pertenece a la asociación')
        elif role == 'administrador_conjunto' and conjunto.pk != user.userprofile.conjunto_id: fila_errores.append('conjunto fuera de tu ámbito')
        if not bloque: fila_errores.append(f'bloque "{bloque_nombre}" no pertenece al conjunto')
        if fila_errores: errores.append(f'Fila {fila}: ' + '; '.join(fila_errores) + '.')
    return errores


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def importar_socios_xlsx_preview(request):
    if request.method != 'POST':
        return redirect('socios:importar_socios_masivo')

    f = request.FILES.get('file')
    if not f:
        messages.error(request, 'Sube un archivo .xlsx.')
        return redirect('socios:importar_socios_masivo')

    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            messages.error(request, 'El archivo estÃ¡ vacÃ­o.')
            return redirect('socios:importar_socios_masivo')

        headers = [str(cell or '').strip() for cell in rows[0]]
        required_headers = ['username', 'nombre', 'apellido_paterno', 'email', 'asociacion', 'conjunto', 'bloque']
        missing_headers = [header for header in required_headers if header not in headers]
        if missing_headers:
            messages.error(request, f'Faltan columnas obligatorias: {", ".join(missing_headers)}.')
            return redirect('socios:importar_socios_masivo')
        preview = [[str(cell or '') for cell in row] for row in rows[1:11]]
        
        def convert_cell_value(cell):
            if cell is None:
                return ''
            if isinstance(cell, datetime):
                return cell.strftime('%Y-%m-%d') if cell.time() == datetime.min.time() else cell.strftime('%Y-%m-%d %H:%M:%S')
            return str(cell)
        
        preview_data = [
            [convert_cell_value(cell) for cell in row[:17]]
            for row in rows[1:]
            if any(cell is not None for cell in row[:17])
        ]

        request.session['socios_import_preview'] = preview_data

        return render(request, 'socios/importar_masivo.html', {
            'preview_headers': headers,
            'preview_rows': preview,
            'import_errors': validar_filas_importacion(preview_data, request.user),
        })
    except Exception as e:
        messages.error(request, f'Error al procesar xlsx: {e}')
        return redirect('socios:importar_socios_masivo')


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def importar_socios_xlsx_confirm(request):
    preview_data = request.session.pop('socios_import_preview', None)
    if not preview_data:
        messages.error(request, 'No hay datos para confirmar.')
        return redirect('socios:importar_socios_masivo')

    import_errors = validar_filas_importacion(preview_data, request.user)
    if import_errors:
        messages.error(request, 'No se registró ningún socio: ' + ' '.join(import_errors[:8]))
        return redirect('socios:importar_socios_masivo')

    created = 0
    skipped = 0
    errors = []
    for row in preview_data:
        vals = [(c or '') for c in (list(row) + [''] * max(0, 17 - len(row)))[:17]]
        username = vals[0]
        nombre = vals[1]
        apellido_paterno = vals[2]
        apellido_materno = vals[3]
        apellido = vals[4] or f"{apellido_paterno} {apellido_materno}".strip()
        email = vals[5]
        password = vals[6]
        telefono = vals[7]
        ciudad = vals[8]
        direccion = vals[9]
        fecha_nacimiento = vals[10] or None
        razon = vals[11]
        carnet_ci = vals[12]
        carnet_complemento = vals[13]
        asociacion = Asociacion.objects.get(nombre__iexact=str(vals[14]).strip(), activo=True)
        conjunto = Conjunto.objects.get(nombre__iexact=str(vals[15]).strip(), asociacion=asociacion, activo=True)
        bloque = Bloque.objects.get(nombre__iexact=str(vals[16]).strip(), conjunto=conjunto, activo=True)
        
        if not username:
            skipped += 1
            errors.append(f"Fila sin username: {nombre}")
            continue
        
        if not password:
            password = User.objects.make_random_password()
        
        try:
            if User.objects.filter(username=username).exists():
                # Actualizar usuario existente
                user = User.objects.get(username=username)
                user.email = email
                if password and password != '':
                    user.set_password(password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                
                # Actualizar o crear socio
                socio, created_socio = Socio.objects.update_or_create(
                    user=user,
                    defaults={
                        'nombre': nombre,
                        'apellido_paterno': apellido_paterno,
                        'apellido_materno': apellido_materno,
                        'apellido': apellido,
                        'email': email,
                        'telefono': telefono,
                        'ciudad': ciudad,
                        'direccion': direccion,
                        'fecha_nacimiento': fecha_nacimiento,
                        'razon': razon,
                        'carnet_ci': carnet_ci,
                        'carnet_complemento': carnet_complemento,
                    }
                )
                if created_socio:
                    created += 1
                else:
                    created += 1  # Contar como actualizado
            else:
                # Crear nuevo usuario
                user = User.objects.create_user(username=username, email=email, password=password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                socio = Socio.objects.create(user=user, codigo_socio=generar_codigo_socio(), nombre=nombre, apellido_paterno=apellido_paterno, apellido_materno=apellido_materno, apellido=apellido, email=email, telefono=telefono, ciudad=ciudad, direccion=direccion, fecha_nacimiento=fecha_nacimiento, razon=razon, carnet_ci=carnet_ci, carnet_complemento=carnet_complemento, creado_por=request.user)
                created += 1
            membresia = socio.membresias.filter(estado__in=['activo', 'suspendido', 'castigado']).first()
            if membresia:
                membresia.bloque = bloque
                membresia.save(update_fields=['bloque'])
            else:
                Membresia.inscribir(socio, asociacion, conjunto, bloque, estado_pago='al_dia')
        except Exception as e:
            skipped += 1
            errors.append(f"Error creando {username}: {str(e)}")

    if errors:
        messages.error(request, f'Socios importados: {created}, omitidos: {skipped}. Errores: {"; ".join(errors[:5])}')
    else:
        messages.success(request, f'Socios importados desde XLSX: {created}')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def descargar_plantilla_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'socios'
    
    # Encabezados
    headers = ['username', 'nombre', 'apellido_paterno', 'apellido_materno', 'apellido', 'email', 'password', 'telefono', 'ciudad', 'direccion', 'fecha_nacimiento', 'razon', 'carnet_ci', 'carnet_complemento', 'asociacion', 'conjunto', 'bloque']
    ws.append(headers)
    
    # Estilo para el encabezado
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Azul
    header_font = Font(color="FFFFFF", bold=True, size=11)  # Letra blanca y negrita
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    column_widths = {
        'A': 15,  # username
        'B': 20,  # nombre
        'C': 20,  # apellido_paterno
        'D': 20,  # apellido_materno
        'E': 25,  # apellido
        'F': 25,  # email
        'G': 15,  # password
        'H': 15,  # telefono
        'I': 15,  # ciudad
        'J': 30,  # direccion
        'K': 15,  # fecha_nacimiento
        'L': 30,  # razon
        'M': 15,  # carnet_ci
        'N': 15,  # carnet_complemento
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # AÃ±adir bordes a todas las celdas
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Ejemplo de fila con estilo
    example_row = ['jdoe', 'Juan', 'Perez', 'Gomez', 'Perez Gomez', 'jdoe@example.com', 'Passw0rd!', '71234567', 'Oruro', 'DirecciÃ³n 123', '1990-01-01', 'Quiero participar', '1234567', '-1A', 'Nombre exacto de asociación', 'Nombre exacto de conjunto', 'Nombre exacto de bloque']
    ws.append(example_row)
    
    # Aplicar bordes y colores alternados a las filas de datos
    for row_num in range(2, ws.max_row + 1):
        # Color alternado para filas
        if row_num % 2 == 0:
            row_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # Azul claro
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Congelar la primera fila (encabezado)
    ws.freeze_panes = "A2"
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=socios_plantilla.xlsx'
    wb.save(response)
    return response


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def importar_socios_xlsx(request):
    if request.method == 'POST':
        f = request.FILES.get('file')
        if not f:
            messages.error(request, 'Sube un archivo .xlsx')
            return redirect('socios:listar_socios')
        try:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            created = 0
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                vals = [ (c or '') for c in row[:14] ]
                username = vals[0]
                nombre = vals[1]
                apellido_paterno = vals[2]
                apellido_materno = vals[3]
                apellido = vals[4] or f"{apellido_paterno} {apellido_materno}".strip()
                email = vals[5]
                password = vals[6]
                telefono = vals[7]
                ciudad = vals[8]
                direccion = vals[9]
                fecha_nacimiento = vals[10] or None
                razon = vals[11]
                carnet_ci = vals[12]
                carnet_complemento = vals[13]
                if not username or User.objects.filter(username=username).exists():
                    continue
                if not password:
                    password = User.objects.make_random_password()
                user = User.objects.create_user(username=username, email=email, password=password)
                user.first_name = nombre
                user.last_name = apellido_paterno or apellido
                user.save()
                Socio.objects.create(user=user, codigo_socio=generar_codigo_socio(), nombre=nombre, apellido_paterno=apellido_paterno, apellido_materno=apellido_materno, apellido=apellido, email=email, telefono=telefono, ciudad=ciudad, direccion=direccion, fecha_nacimiento=fecha_nacimiento, razon=razon, carnet_ci=carnet_ci, carnet_complemento=carnet_complemento)
                created += 1
            messages.success(request, f'Socios importados desde XLSX: {created}')
        except Exception as e:
            messages.error(request, f'Error al procesar xlsx: {e}')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(can_manage_users, login_url='/login/')
def listar_admins(request):
    q = request.GET.get('q', '').strip()
    activo = request.GET.get('activo', '').strip()
    admins = User.objects.filter(is_staff=True).select_related('userprofile', 'userprofile__asociacion', 'userprofile__conjunto')

    if q:
        admins = admins.filter(
            Q(username__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
        )

    if activo == 'si':
        admins = admins.filter(is_active=True)
    elif activo == 'no':
        admins = admins.filter(is_active=False)

    paginator = Paginator(admins.order_by('username'), 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'admins/admins.html', {
        'page_obj': page_obj,
        'q': q,
        'activo': activo,
        'asociaciones': Asociacion.objects.filter(activo=True),
        'conjuntos': Conjunto.objects.filter(activo=True).select_related('asociacion'),
    })


@login_required
@user_passes_test(can_manage_users, login_url='/login/')
def ver_admin(request, user_id):
    return redirect('socios:listar_admins')


@login_required
@user_passes_test(can_manage_users, login_url='/login/')
def editar_admin(request, user_id):
    user = get_object_or_404(User, id=user_id, is_staff=True)
    if request.method == 'POST':
        user.username = request.POST.get('username', user.username).strip()
        user.email = request.POST.get('email', user.email).strip()
        user.first_name = request.POST.get('first_name', user.first_name).strip()
        user.last_name = request.POST.get('last_name', user.last_name).strip()
        rol = request.POST.get('rol', 'administrador_asociacion')
        asociacion_id = request.POST.get('asociacion_id') or None
        asociacion = Asociacion.objects.filter(pk=asociacion_id, activo=True).first() if asociacion_id else None
        conjunto_id = request.POST.get('conjunto_id') or None
        conjunto = Conjunto.objects.filter(pk=conjunto_id, asociacion=asociacion, activo=True).first() if asociacion and conjunto_id else None
        if rol not in {'superadministrador', 'administrador_asociacion', 'administrador_conjunto'}:
            messages.error(request, 'Selecciona un tipo de administrador válido.')
            return redirect('socios:listar_admins')
        if (rol == 'administrador_asociacion' and not asociacion) or (rol == 'administrador_conjunto' and (not asociacion or not conjunto)):
            messages.error(request, 'El tipo de administrador requiere un ámbito válido.')
            return redirect('socios:listar_admins')
        password = request.POST.get('password')
        if password:
            user.set_password(password)
        user.save()
        UserProfile.objects.filter(user=user).update(rol=rol, asociacion=asociacion, conjunto=conjunto)
        registrar_auditoria(
            request.user,
            'modificacion_admin',
            f'Admin {user.username} ({rol})',
            nuevo={'username': user.username, 'rol': rol, 'asociacion': asociacion.nombre if asociacion else None, 'conjunto': conjunto.nombre if conjunto else None},
            asociacion=asociacion,
            conjunto=conjunto,
        )
        messages.success(request, 'Administrador actualizado.')
        return redirect('socios:listar_admins')
    return redirect('socios:listar_admins')


@login_required
@user_passes_test(can_manage_users, login_url='/login/')
def eliminar_admin(request, user_id):
    user = get_object_or_404(User, id=user_id, is_staff=True)
    if request.method == 'POST':
        username = user.username
        user.delete()
        registrar_auditoria(request.user, 'eliminacion_admin', f'Admin {username}')
        messages.success(request, 'Administrador eliminado.')
        return redirect('socios:listar_admins')
    return redirect('socios:listar_admins')


@login_required
def mis_souvenirs(request):
    try:
        socio = request.user.socio_profile
    except Socio.DoesNotExist:
        messages.error(request, 'No se encontró perfil de socio.')
        return redirect('/')
    entregas = socio.entregas_souvenir.select_related('souvenir', 'entregado_por').all()
    paginator = Paginator(entregas, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'socios/mis_souvenirs.html', {'page_obj': page_obj})


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def editar_socio(request, socio_id):
    if request.method != 'POST':
        return redirect('socios:listar_socios')

    socio = get_object_or_404(scope_socios(Socio.objects.all(), request.user), id=socio_id)
    membresia = socio.membresias.filter(estado__in=['activo', 'suspendido', 'castigado']).first() or socio.membresias.order_by('-fecha_ingreso').first()
    socio.nombre = request.POST.get('nombre', '').strip()
    socio.apellido_paterno = request.POST.get('apellido_paterno', '').strip()
    socio.apellido_materno = request.POST.get('apellido_materno', '').strip()
    socio.apellido = request.POST.get('apellido', '').strip() or f"{socio.apellido_paterno} {socio.apellido_materno}".strip()
    socio.email = request.POST.get('email', '').strip()
    socio.telefono = request.POST.get('telefono', '').strip()
    socio.ciudad = request.POST.get('ciudad', '').strip()
    socio.direccion = request.POST.get('direccion', '').strip()
    socio.carnet_ci = request.POST.get('carnet_ci', '').strip()
    socio.carnet_complemento = request.POST.get('carnet_complemento', '').strip()
    socio.observacion = request.POST.get('observacion', '').strip()
    socio.fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
    socio.sexo = request.POST.get('sexo', '').strip()
    socio.modalidad = request.POST.get('modalidad', '').strip()
    if membresia:
        role = get_role(request.user)
        conjunto = membresia.conjunto
        if role == 'superadministrador':
            asociacion = Asociacion.objects.filter(pk=request.POST.get('asociacion_id'), activo=True).first()
            conjunto = Conjunto.objects.filter(pk=request.POST.get('conjunto_id'), asociacion=asociacion, activo=True).first() if asociacion else None
        elif role == 'administrador_asociacion':
            conjunto = Conjunto.objects.filter(pk=request.POST.get('conjunto_id'), asociacion_id=request.user.userprofile.asociacion_id, activo=True).first()
        bloque_id = request.POST.get('bloque_id')
        bloque = Bloque.objects.filter(pk=bloque_id, conjunto=conjunto, activo=True).first() if conjunto else None
        if not bloque:
            messages.error(request, 'Selecciona una asociación, conjunto y bloque válidos para el socio.')
            return redirect('socios:listar_socios')
        membresia.asociacion = conjunto.asociacion
        membresia.conjunto = conjunto
        membresia.bloque = bloque
        membresia.save(update_fields=['asociacion', 'conjunto', 'bloque'])
    socio.save()
    registrar_auditoria(request.user, 'modificacion_socio', f'Socio {socio.pk} - {socio}', nuevo={'nombre': socio.nombre, 'email': socio.email})
    messages.success(request, 'Datos del socio actualizados.')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def activar_socio(request, socio_id):
    socio = get_object_or_404(scope_socios(Socio.objects.all(), request.user), id=socio_id)
    anteriores = list(socio.membresias.filter(estado__in=['suspendido', 'castigado']).values_list('id', flat=True))
    socio.membresias.filter(estado__in=['suspendido', 'castigado']).update(estado='activo')
    registrar_auditoria(request.user, 'activacion_membresia', f'Socio {socio.pk} - {socio}', anterior={'membresias': anteriores}, nuevo={'estado': 'activo'})
    messages.success(request, 'Membresía activa.')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(can_register_members, login_url='/login/')
def desactivar_socio(request, socio_id):
    socio = get_object_or_404(scope_socios(Socio.objects.all(), request.user), id=socio_id)
    anteriores = list(socio.membresias.values_list('id', flat=True))
    socio.membresias.update(estado='baja')
    registrar_auditoria(request.user, 'baja_membresia', f'Socio {socio.pk} - {socio}', anterior={'membresias': anteriores}, nuevo={'estado': 'baja'})
    messages.success(request, 'Membresía dada de baja.')
    return redirect('socios:listar_socios')


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def historial_souvenirs(request, socio_id):
    socio = get_object_or_404(scope_socios(Socio.objects.all(), request.user), id=socio_id)
    entregas = socio.entregas_souvenir.select_related('souvenir', 'evento', 'entregado_por').order_by('-fecha_entrega')
    total_entregas = entregas.count()
    paginator = Paginator(entregas, 10)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'socios/historial_souvenirs.html', {
        'socio': socio,
        'page_obj': page_obj,
        'total_entregas': total_entregas,
    })


@login_required
@user_passes_test(is_administrative, login_url='/login/')
def eliminar_socio(request, socio_id):
    socio = get_object_or_404(scope_socios(Socio.objects.all(), request.user), id=socio_id)
    socio_info = f'Socio {socio.pk} - {socio}'
    if getattr(socio, 'user', None):
        socio.user.delete()
    else:
        socio.delete()
    registrar_auditoria(request.user, 'eliminacion_socio', socio_info)
    messages.success(request, 'Socio eliminado definitivamente.')
    return redirect('socios:listar_socios')


